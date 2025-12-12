"""Excel file viewer."""

from textual import work
from textual.app import ComposeResult
from textual.widgets import DataTable, Static, TabbedContent, TabPane

from quickview.viewers.base import BaseViewer


class ExcelSheetWidget(DataTable):
    """Widget to display a single Excel sheet."""

    def __init__(self, filepath: str, sheet_name: str, **kwargs):
        super().__init__(**kwargs)
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.cursor_type = "row"
        self.zebra_stripes = True

    def on_mount(self) -> None:
        self._load_sheet()

    @work(thread=True)
    def _load_sheet(self) -> None:
        try:
            import openpyxl
        except ImportError:
            return

        try:
            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            sheet = wb[self.sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            wb.close()

            if rows:
                self.app.call_from_thread(self._populate_table, rows)

        except Exception as e:
            self.app.call_from_thread(
                self.app.notify, f"Error loading sheet {self.sheet_name}: {e}", severity="error"
            )

    def _populate_table(self, rows: list) -> None:
        if not rows:
            return

        # Filter out completely empty rows
        rows = [row for row in rows if row and any(cell is not None for cell in row)]
        if not rows:
            return

        max_cols = max(len(row) for row in rows)

        def col_name(n: int) -> str:
            result = ""
            while n >= 0:
                result = chr(65 + (n % 26)) + result
                n = n // 26 - 1
            return result

        for i in range(max_cols):
            self.add_column(col_name(i), key=f"col_{i}")

        for row in rows:
            str_row = [str(cell) if cell is not None else "" for cell in row]
            while len(str_row) < max_cols:
                str_row.append("")
            self.add_row(*str_row[:max_cols])


class ExcelViewer(BaseViewer):
    """Viewer for Excel files."""

    extensions = [".xlsx", ".xls", ".xlsm"]

    @staticmethod
    def _sanitize_id(name: str) -> str:
        """Convert sheet name to valid CSS identifier."""
        import re
        # Replace non-alphanumeric with underscore, ensure starts with letter
        sanitized = re.sub(r'[^a-zA-Z0-9]', '_', name)
        if sanitized and sanitized[0].isdigit():
            sanitized = 's' + sanitized
        return sanitized or 'sheet'

    def compose(self) -> ComposeResult:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            if len(sheet_names) == 1:
                safe_id = self._sanitize_id(sheet_names[0])
                yield ExcelSheetWidget(
                    self.filepath, sheet_names[0],
                    id=f"table-{safe_id}"
                )
            else:
                with TabbedContent():
                    for i, sheet_name in enumerate(sheet_names):
                        safe_id = f"{i}_{self._sanitize_id(sheet_name)}"
                        with TabPane(sheet_name, id=f"sheet-{safe_id}"):
                            yield ExcelSheetWidget(
                                self.filepath, sheet_name,
                                id=f"table-{safe_id}", classes="excel-table"
                            )

        except ImportError:
            yield Static("[red]Error: openpyxl not installed. Run: pip install quickview[excel][/red]")
        except Exception as e:
            yield Static(f"[red]Error loading Excel file: {e}[/red]")

    def load(self) -> None:
        # Loading is handled by ExcelSheetWidget.on_mount
        pass
